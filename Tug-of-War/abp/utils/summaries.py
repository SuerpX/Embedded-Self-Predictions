import os
import logging
import shutil
from openpyxl import Workbook


def clear_summary_path(path_to_summary):
    """ Removes the summaries if it exists """
    if os.path.exists(path_to_summary):
        logging.info("Summaries Exists. Deleting the summaries at %s" % path_to_summary)
        shutil.rmtree(path_to_summary)


def saliency_to_excel(saliencies, choices, reward_types, choice_descriptions,
                      layers, game_info, dest_filename='explanation.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Game Info"
    for key, value in game_info.items():
        row = ws.max_row + 1
        ws['A' + str(row)] = key
        ws['B' + str(row)] = value

    for choice_idx, choice in enumerate(choices):
        action = choice_descriptions[choice_idx]
        ws = wb.create_sheet(title=action)
        saliency = saliencies[choice]["all"]
        for idx, layer in enumerate(layers):
            current_row = ws.max_row + 1
            ws['A' + str(current_row)] = layer
            for layer_row in saliency[:, :, idx]:
                ws.append(layer_row.tolist())

        for reward_index, reward_type in enumerate(reward_types):
            key = "{}({})".format(action, reward_type)
            ws = wb.create_sheet(title=key)
            saliency = saliencies[choice][reward_type]
            for idx, layer in enumerate(layers):
                current_row = ws.max_row + 1
                ws['A' + str(current_row)] = layer
                for layer_row in saliency[:, :, idx]:
                    ws.append(layer_row.tolist())

    wb.save(filename=dest_filename)
